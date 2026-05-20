#!/usr/bin/env python3
import sys
import json
import openpyxl
from io import BytesIO

def parse_excel_all_sheets(buffer: bytes) -> list:
    wb = openpyxl.load_workbook(BytesIO(buffer), data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        headers = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val) if val is not None else f'column_{len(headers)}')

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            row_obj = {}
            for col_idx, val in enumerate(row):
                field = headers[col_idx] if col_idx < len(headers) else f'column_{col_idx}'
                row_obj[field] = val if val is not None else ''
            rows.append(row_obj)

        if rows:
            sheets.append({
                'sheetName': sheet_name,
                'fields': headers,
                'data': rows,
                'rowCount': len(rows)
            })

    return sheets

if __name__ == '__main__':
    buffer = sys.stdin.buffer.read()
    sheets = parse_excel_all_sheets(buffer)
    print(json.dumps(sheets, ensure_ascii=False))